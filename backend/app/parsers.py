from io import BytesIO

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def parse_docx(data: bytes) -> str:
    doc = Document(BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_pdf(data: bytes) -> str:
    reader = PdfReader(BytesIO(data))
    pages = [p.extract_text() or "" for p in reader.pages]
    return "\n".join(pages)


def parse_xlsx(data: bytes) -> str:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    rows: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals = [str(cell).strip() for cell in row if cell not in (None, "")]
            if vals:
                rows.append(" | ".join(vals))
    return "\n".join(rows)


def parse_by_extension(filename: str, data: bytes) -> str:
    name = filename.lower()
    if name.endswith(".docx"):
        return parse_docx(data)
    if name.endswith(".xlsx"):
        return parse_xlsx(data)
    if name.endswith(".pdf"):
        return parse_pdf(data)
    raise ValueError(f"Unsupported file type: {filename}")
